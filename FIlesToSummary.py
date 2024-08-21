import glob
import os
from tkinter import Tk
from tkinter.filedialog import askdirectory, askopenfilename
import pandas as pd
from openpyxl import load_workbook


# Function to find the first empty row in a sheet


def find_first_empty_row(sheet):
    col = sheet['A']
    for idx, cell in enumerate(col, start=1):
        if cell.value is None:
            return idx
    return len(col) + 1


# Function to find the last row with data in the DataFrame
def extract_last_row_cells(kgp):
    # Read the Excel file with header starting from the 5th row only for this function
    df = pd.read_excel(kgp, header=4)

    # Drop any rows that are completely empty
    df = df.dropna(how='all')

    # Check if there are any remaining rows after dropping empty rows
    if df.empty:
        print("The DataFrame is empty after dropping all-empty rows.")
        return None

    # Print the DataFrame after dropping empty rows
    print("DataFrame after dropping empty rows:")
    print(df)

    # Find the last row with data based on specific columns
    columns_to_check = ['ΚΟΛΛΑ', 'TEMAXIA', 'Χ.Μ.', 'Χ.Κ.', 'ΑΞΙΑ $', 'ΝΑΥΛΟΣ €']
    for idx in reversed(df.index):
        row = df.loc[idx, columns_to_check]
        if row.notna().any():  # Check if any value in the row is not NaN
            last_row_idx = idx
            last_row = row
            break

    # Print the last row found
    print(f"Last row found at index {last_row_idx}:")
    print(last_row)

    # Extract the specific cells from the last row
    last_row_data = [
        last_row.get('ΚΟΛΛΑ', ''),
        last_row.get('TEMAXIA', ''),
        last_row.get('Χ.Μ.', ''),
        last_row.get('Χ.Κ.', ''),
        last_row.get('ΑΞΙΑ $', ''),
        last_row.get('ΝΑΥΛΟΣ €', '')
    ]

    # Print the extracted data before handling missing values
    print("Extracted data before handling missing values:")
    for col_name, value in zip(columns_to_check, last_row_data):
        print(f"Column '{col_name}': Value '{value}'")

    # Handle missing values by replacing NaN with a default value, e.g., '-'
    last_row_data = [str(value) if pd.notna(value) else '-' for value in last_row_data]

    # Print the final extracted data
    print("Final extracted data:")
    for col_name, value in zip(columns_to_check, last_row_data):
        print(f"Column '{col_name}': Value '{value}'")

    return last_row_data


# Function to print the required lines for each file into the summary Excel file
def print_required_lines(ws, start_row, file, hs_code, last_row_cells_def):
    file_name = os.path.basename(file).split('.')[0]
    folder_name = os.path.basename(os.path.dirname(file))
    file_name_with_folder = f"{file_name}/{folder_name}"

    # First line format
    ws.cell(row=start_row, column=1, value=file_name_with_folder)
    ws.cell(row=start_row, column=2, value=hs_code)
    ws.cell(row=start_row, column=3, value="FZ6159")
    ws.cell(row=start_row, column=4, value="CHINA")
    for idx, cell_value in enumerate(last_row_cells_def, start=5):
        ws.cell(row=start_row, column=idx, value=cell_value)

    # Second line format
    ws.cell(row=start_row + 1, column=2, value=str(hs_code) + " Total")
    for idx, cell_value in enumerate(last_row_cells_def, start=5):
        ws.cell(row=start_row + 1, column=idx, value=cell_value)

    # Third line format
    ws.cell(row=start_row + 2, column=1, value=file_name_with_folder + " Total")
    for idx, cell_value in enumerate(last_row_cells_def, start=5):
        ws.cell(row=start_row + 2, column=idx, value=cell_value)

    # Increment start_row for the next entry
    return start_row + 3  # Move to the next set of three lines


def main():
    # Hide the main Tkinter window
    Tk().withdraw()

    # Open a dialog box to select the directory containing the Excel files
    input_dir = askdirectory(title="Select the directory containing the Excel files")

    # Open a dialog box to select the existing 'sygkentrwtiko.xlsx' file
    summary_path = askopenfilename(title="Select the summary Excel file (sygkentrwtiko.xlsx)",
                                   filetypes=[("Excel files", "*.xlsx")])

    # Load the existing 'sygkentrwtiko.xlsx' file
    wb = load_workbook(summary_path)
    ws = wb.active

    # Find the first empty row in column A, starting at a minimum of row 3
    start_row = find_first_empty_row(ws)

    # Process each file
    for file in glob.glob(os.path.join(input_dir, "*.xlsx")):
        try:
            # Extract necessary data from the file
            last_row_cells_main = extract_last_row_cells(file)

            if last_row_cells_main is None:
                print(f"No valid data found in {file}. Skipping this file.")
                continue

            # Read the Excel file again, now with default header behavior for other operations
            df = pd.read_excel(file)

            # Check if 'HS CODE' column exists after stripping whitespace
            df.columns = df.columns.str.strip()  # Strip whitespace from column headers
            if 'HS CODE' not in df.columns:
                print(f"Column 'HS CODE' not found in {file}. Skipping this file.")
                continue

            # Extract necessary data from the file
            hs_code = df.iloc[0]['HS CODE']  # Adjust this based on your actual column name or index

            # Debugging: Print extracted HS CODE
            print(f"Extracted HS CODE from {file}: {hs_code}")

            # Print required lines for each file into the summary Excel file
            start_row = print_required_lines(ws, start_row, file, hs_code, last_row_cells_main)

        except Exception as e:
            print(f"An error occurred while processing {file}: {e}")
            continue

    ws.cell(row=start_row, column=1, value='Grand Total')
    # Save the updated 'sygkentrwtiko.xlsx' file
    wb.save(summary_path)

    print(f"Data combined successfully into {summary_path}")


if __name__ == "__main__":
    main()
import openpyxl
import tkinter as tk
from tkinter import filedialog


def find_first_empty_cell_in_column(ws, column, start_row):
    """Find the first empty cell in a specific column starting from a given row"""
    for row in range(start_row, ws.max_row):
        if ws.cell(row=row, column=column).value is None:
            return row
    return ws.max_row + 1


def update_files_with_column_data(main_file_path_def, target_dir_def):
    try:
        # Open the main workbook
        main_workbook = openpyxl.load_workbook(main_file_path_def)
        main_worksheet = main_workbook.active  # Assuming data starts from the first sheet

        # Get the number of rows with data in the main worksheet
        row_count = main_worksheet.max_row

        for i in range(2, row_count + 1):
            # Read the filename from column G (7) and data from columns C (3), D (4), and E (5)
            file_name = main_worksheet.cell(row=i, column=7).value
            data_c = main_worksheet.cell(row=i, column=3).value
            data_d = main_worksheet.cell(row=i, column=4).value
            data_e = main_worksheet.cell(row=i, column=5).value

            if file_name:
                target_file_path = f"{target_dir_def}/{file_name}.xlsx"

                try:
                    # Open the target workbook
                    target_workbook = openpyxl.load_workbook(target_file_path)
                    target_worksheet = target_workbook.active  # Assuming data is on the first sheet

                    # Find the first empty cell in column C after the 5th row in the target worksheet
                    target_row = find_first_empty_cell_in_column(target_worksheet, column=3, start_row=6)

                    # Write the data from C, D, and E into the new rows in the target worksheet
                    target_worksheet.cell(row=target_row, column=3, value=data_c)
                    target_worksheet.cell(row=target_row, column=4, value=data_d)
                    target_worksheet.cell(row=target_row, column=7, value=data_e)

                    # Save the target workbook
                    target_workbook.save(target_file_path)

                except FileNotFoundError:
                    print(f"Error: Unable to open the target workbook '{file_name}'.")

    except FileNotFoundError:
        print("Error: Unable to open the main workbook.")


def select_main_file():
    main_file_path_main = filedialog.askopenfilename(title="Select Main Excel File", filetypes=[("Excel files",
                                                                                                 "*.xlsx *.xls")])
    return main_file_path_main


def select_target_directory():
    target_dir_main = filedialog.askdirectory(title="Select Target Directory")
    return target_dir_main


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    main_file_path = select_main_file()
    if not main_file_path:
        print("No main file selected. Exiting.")
        exit()

    target_dir = select_target_directory()
    if not target_dir:
        print("No target directory selected. Exiting.")
        exit()

    update_files_with_column_data(main_file_path, target_dir)
